"""
pipeline.py — toàn bộ logic xử lý, không phụ thuộc PyQt.

Có thể import và dùng độc lập từ dòng lệnh, tiện cho việc test
mà không cần mở UI.
"""

from __future__ import annotations

import base64
import csv
import difflib
import io
import json
import re
import shutil
import threading
import unicodedata
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path

import fitz  # PyMuPDF
from PIL import Image

# ============================================================ CẤU HÌNH

DEFAULT_CONFIG_FILE = "config.json"

PAGE_STRATEGIES = {
    "first": "đầu",
    "first_last": "đầu + cuối",
    "full": "full",
}

EMPTY_VALUES = {
    "", "null", "none", "n/a", "na", "-", "parse_error",
    "không có", "khong co", "không xác định", "khong xac dinh",
    "không rõ", "khong ro", "chưa xác định", "chua xac dinh",
}

GEO_MARKERS = [
    "tỉnh", "thành phố", "tp.", "tp", "huyện", "quận",
    "xã", "phường", "thị xã", "thị trấn", "đặc khu",
]

# Liệt kê tường minh chữ hoa tiếng Việt. KHÔNG dùng dải [A-ZĐÀ-Ỹ]:
# dải À-Ỹ (U+00C0..U+1EF8) chứa cả chữ thường Latin-1 (à-ÿ) nên nó sẽ
# khớp sai. Đây là loại lỗi im lặng rất khó phát hiện.
VN_UPPER = (
    "AĂÂBCDĐEÊGHIKLMNOÔƠPQRSTUƯVXY"
    "ÁÀẢÃẠẮẰẲẴẶẤẦẨẪẬ"
    "ÉÈẺẼẸẾỀỂỄỆ"
    "ÍÌỈĨỊ"
    "ÓÒỎÕỌỐỒỔỖỘỚỜỞỠỢ"
    "ÚÙỦŨỤỨỪỬỮỰ"
    "ÝỲỶỸỴ"
)

_PLACEHOLDER = "\u27e6P{}\u27e7"

# Hàng rào cho bước sửa lỗi
MIN_SIMILARITY = 0.80
MAX_LENGTH_DRIFT = 0.25
CORRECTION_MAX_TOKENS = 160
# ProtonX chạy trên server nên không còn tokenizer để đo độ dài; chia câu
# dài theo số ký tự (≈ 160 token seq2seq).
CORRECTION_MAX_CHARS = 600


def load_config(path=DEFAULT_CONFIG_FILE) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_fields(path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_fields(path, fields: dict) -> Path | None:
    """Ghi lại data_fields.json, có sao lưu. Trả về đường dẫn file .bak."""
    path = Path(path)
    backup = None

    if path.exists():
        stamp = datetime.now().strftime("%d%m%Y-%H%M%S")
        backup = path.with_suffix(f".{stamp}.bak")
        shutil.copy2(path, backup)

    clean = {
        k: unicodedata.normalize("NFC", str(v)).strip()
        for k, v in fields.items()
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(clean, f, ensure_ascii=False, indent=2)

    return backup


# ============================================================ QUÉT NGUỒN

def collect_pdf_files(source: str, recursive=True) -> list[Path]:
    """Nhận file .pdf hoặc thư mục, trả về danh sách file đã sắp xếp."""
    p = Path(source)
    if p.is_file():
        return [p] if p.suffix.lower() == ".pdf" else []
    if not p.is_dir():
        return []
    pattern = "**/*.pdf" if recursive else "*.pdf"
    return sorted(f for f in p.glob(pattern) if f.is_file())


def pdf_page_count(pdf_path) -> int:
    try:
        with fitz.open(pdf_path) as doc:
            return doc.page_count
    except Exception:
        return 0


def select_pages(total_pages: int, strategy: str) -> list[int]:
    """Chỉ số trang 0-based cần OCR."""
    if total_pages <= 0:
        return []
    if total_pages == 1 or strategy == "first":
        return [0]
    if strategy == "first_last":
        return [0, total_pages - 1]
    return list(range(total_pages))          # full


def count_selected_pages(total_pages: int, strategy: str) -> int:
    """Số trang sẽ OCR — tính trực tiếp, KHÔNG dựng list. Quan trọng khi ước
    lượng thư mục lớn: 'full' trên PDF hàng nghìn trang không tạo range khổng lồ."""
    if total_pages <= 0:
        return 0
    if total_pages == 1 or strategy == "first":
        return 1
    if strategy == "first_last":
        return 2
    return total_pages                        # full


def estimate_workload(files, strategy, cancel_check=None) -> tuple[int, int]:
    """Trả về (số_file, tổng_số_trang_sẽ_OCR). Nhanh vì không render.
    cancel_check: callable trả True để dừng sớm (thư mục rất lớn)."""
    pages = 0
    for f in files:
        if cancel_check and cancel_check():
            break
        pages += count_selected_pages(pdf_page_count(f), strategy)
    return len(files), pages


# ============================================================ RENDER

def render_pages(pdf_path, page_numbers, dpi, max_dimension, quality):
    """Trả về list (page_index, data_uri). Mở PDF một lần cho mọi trang."""
    out = []
    with fitz.open(pdf_path) as doc:
        total = doc.page_count
        mat = fitz.Matrix(dpi / 72, dpi / 72)

        for pno in page_numbers:
            if pno < 0 or pno >= total:
                continue

            pix = doc[pno].get_pixmap(matrix=mat)
            img = Image.open(io.BytesIO(pix.tobytes("png")))
            if img.mode != "RGB":
                img = img.convert("RGB")

            if img.width > max_dimension or img.height > max_dimension:
                ratio = min(max_dimension / img.width,
                            max_dimension / img.height)
                img = img.resize(
                    (int(img.width * ratio), int(img.height * ratio)),
                    Image.Resampling.LANCZOS,
                )

            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=quality, optimize=True)
            b64 = base64.b64encode(buf.getvalue()).decode()
            out.append((pno, f"data:image/jpeg;base64,{b64}"))
    return out


# ============================================================ GỌI MODEL

def build_prompt(fields: dict, page_index: int, total_pages: int) -> str:
    if total_pages == 1:
        ctx = "Đây là văn bản một trang."
    elif page_index == 0:
        ctx = f"Đây là TRANG ĐẦU (trang 1/{total_pages}) của văn bản."
    elif page_index == total_pages - 1:
        ctx = f"Đây là TRANG CUỐI (trang {page_index + 1}/{total_pages}) của văn bản."
    else:
        ctx = f"Đây là trang {page_index + 1}/{total_pages} của văn bản."

    return (
        f"{ctx}\n"
        "Trích xuất các dữ liệu trong ảnh. "
        "Nếu một trường KHÔNG xuất hiện trên trang này, để giá trị là chuỗi rỗng \"\". "
        "TUYỆT ĐỐI không suy đoán hay bịa giá trị.\n"
        "Trả về CHÍNH XÁC dạng JSON theo mẫu (không thêm text giải thích):\n"
        + json.dumps(fields, ensure_ascii=False)
    )


def call_ocr(client, model_name, image_url, prompt, max_tokens):
    resp = client.chat.completions.create(
        model=model_name,
        messages=[{
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": image_url}},
            ],
        }],
        temperature=0,
        top_p=1.0,
        max_tokens=max_tokens,
    )
    return resp.choices[0].message.content


def parse_json_loose(raw):
    if not raw:
        return None
    s = str(raw).strip()
    try:
        if s.startswith("{") and s.endswith("}"):
            return json.loads(s)
        i, j = s.find("{"), s.rfind("}") + 1
        if i != -1 and j > i:
            return json.loads(s[i:j])
    except Exception:
        pass
    return None


# ============================================================ GỘP KẾT QUẢ

def is_empty(value) -> bool:
    if value is None:
        return True
    return str(value).strip().lower() in EMPTY_VALUES


def normalize_vi(value):
    if isinstance(value, str):
        return unicodedata.normalize("NFC", value).strip()
    return value


def merge_page_results(page_results, field_keys, preference: dict):
    """
    page_results: list (page_index, parsed_dict)
    Trả về (merged, provenance) — provenance là số trang 1-based.
    """
    merged = {k: "" for k in field_keys}
    provenance = {k: None for k in field_keys}
    ordered = sorted(page_results, key=lambda x: x[0])

    for key in field_keys:
        pref = preference.get(key, "first")
        candidates = ordered if pref == "first" else list(reversed(ordered))
        for pno, data in candidates:
            if not isinstance(data, dict):
                continue
            val = normalize_vi(data.get(key))
            if not is_empty(val):
                merged[key] = val
                provenance[key] = pno + 1
                break

    return merged, provenance


# ============================================================ SỬA LỖI

_corrector_cache = {}
_corrector_lock = threading.Lock()


class ProtonXCorrector:
    """
    Bọc protonx-legal-tc chạy trên server OpenAI-compatible (không nạp model ở
    máy client, không cần torch). Model là seq2seq sinh văn bản nên MỌI kết quả
    đều qua hàng rào; trượt hàng rào thì giữ nguyên bản gốc.
    """

    def __init__(self, endpoint, model, api_key="EMPTY", num_beams=10,
                 protected_terms=None, use_chat=False, timeout=60):
        from openai import OpenAI

        self.client = OpenAI(api_key=api_key or "EMPTY", base_url=endpoint,
                             timeout=timeout)
        self.model = model
        self.use_chat = use_chat
        self.num_beams = num_beams
        self.device = "server"        # giữ thuộc tính cho phần log/tương thích
        self.protected_terms = list(protected_terms or [])
        self.stats = {"corrected": 0, "unchanged": 0, "rejected": 0, "skipped": 0}

    # ---- gọi model trên server (giải mã tham lam, ổn định qua nhiều bản vLLM)

    def _generate(self, text):
        if self.use_chat:
            resp = self.client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": text}],
                temperature=0,
                max_tokens=CORRECTION_MAX_TOKENS,
            )
            return (resp.choices[0].message.content or "").strip()
        resp = self.client.completions.create(
            model=self.model,
            prompt=text,
            temperature=0,
            max_tokens=CORRECTION_MAX_TOKENS,
        )
        return (resp.choices[0].text or "").strip()

    # ---- hàng rào

    @staticmethod
    def _digits(s):
        return re.findall(r"\d", s or "")

    def _passes_guard(self, original, corrected):
        if not corrected or not corrected.strip():
            return False, "output rỗng"
        if self._digits(original) != self._digits(corrected):
            return False, "chữ số bị thay đổi"
        if original:
            drift = abs(len(corrected) - len(original)) / len(original)
            if drift > MAX_LENGTH_DRIFT:
                return False, f"độ dài lệch {drift:.0%}"
        ratio = difflib.SequenceMatcher(None, original, corrected).ratio()
        if ratio < MIN_SIMILARITY:
            return False, f"tương đồng {ratio:.2f}"
        return True, ""

    # ---- mask danh từ riêng

    def _mask(self, text):
        mapping, masked, idx = {}, text, 0

        for term in sorted(self.protected_terms, key=len, reverse=True):
            if term and term in masked:
                ph = _PLACEHOLDER.format(idx)
                masked = masked.replace(term, ph)
                mapping[ph] = term
                idx += 1

        # Sinh biến thể hoa/thường của từ khoá thay vì dùng re.IGNORECASE.
        # IGNORECASE sẽ lan sang class chữ hoa ở nhóm sau và bắt luôn từ
        # nối chữ thường: "tỉnh Bắc Giang và Sở" bị bắt trọn thay vì "Bắc Giang".
        variants = set()
        for m in GEO_MARKERS:
            variants.update({m, m.capitalize(), m.upper()})
        marker_alt = "|".join(
            re.escape(v) for v in sorted(variants, key=len, reverse=True))

        pattern = re.compile(
            rf"\b({marker_alt})\s+"
            rf"((?:[{VN_UPPER}][\w'\-]*)(?:\s+[{VN_UPPER}][\w'\-]*){{0,3}})",
            flags=re.UNICODE,
        )

        def _sub(m):
            nonlocal idx
            ph = _PLACEHOLDER.format(idx)
            mapping[ph] = m.group(2)
            idx += 1
            return f"{m.group(1)} {ph}"

        return pattern.sub(_sub, masked), mapping

    @staticmethod
    def _unmask(text, mapping):
        for ph, original in mapping.items():
            text = text.replace(ph, original)
        return text

    # ---- chia câu dài

    def _split_if_long(self, text):
        if len(text) <= CORRECTION_MAX_CHARS:
            return [text]

        chunks, buf = [], ""
        for part in re.split(r"(?<=[.;:!?])\s+", text):
            cand = (buf + " " + part).strip() if buf else part
            if len(cand) <= CORRECTION_MAX_CHARS:
                buf = cand
            else:
                if buf:
                    chunks.append(buf)
                buf = part
        if buf:
            chunks.append(buf)
        return chunks or [text]

    # ---- API

    def correct_text(self, text, policy="correct"):
        if policy == "skip" or not text or not str(text).strip():
            self.stats["skipped"] += 1
            return text, "skipped"

        original = unicodedata.normalize("NFC", str(text)).strip()
        working, mapping = (self._mask(original) if policy == "protect"
                            else (original, {}))

        pieces = []
        for chunk in self._split_if_long(working):
            try:
                pieces.append(self._generate(chunk))
            except Exception:
                pieces.append(chunk)

        corrected = unicodedata.normalize("NFC", " ".join(pieces)).strip()
        if mapping:
            corrected = self._unmask(corrected, mapping)

        ok, reason = self._passes_guard(original, corrected)
        if not ok:
            self.stats["rejected"] += 1
            return original, f"rejected:{reason}"
        if corrected == original:
            self.stats["unchanged"] += 1
            return original, "unchanged"

        self.stats["corrected"] += 1
        return corrected, "corrected"

    def correct_fields(self, data: dict, policy_map: dict):
        out, audit = {}, {}
        for key, value in data.items():
            policy = policy_map.get(key, "skip")
            new_value, status = self.correct_text(value, policy)
            out[key] = new_value
            if status == "corrected" or status.startswith("rejected"):
                audit[key] = {"policy": policy, "status": status,
                              "before": value, "after": new_value}
        return out, audit


def get_corrector(endpoint, model, api_key, num_beams, protected_terms,
                  use_chat=False):
    """Cache theo (endpoint, model) — tạo client một lần cho cả phiên."""
    key = (endpoint, model)
    with _corrector_lock:
        if key not in _corrector_cache:
            _corrector_cache[key] = ProtonXCorrector(
                endpoint, model, api_key, num_beams, protected_terms, use_chat)
        return _corrector_cache[key]


# ============================================================ XỬ LÝ 1 FILE

def process_pdf(pdf_path, fields, client, model_name, cfg,
                corrector=None, cancel_check=None):
    """
    Xử lý một PDF. Trả về dict kết quả hoặc raise Exception.

    cancel_check: callable trả True nếu cần dừng ngay giữa file.
    """
    field_keys = list(fields.keys())
    total = pdf_page_count(pdf_path)
    if total == 0:
        raise RuntimeError("không mở được PDF")

    strategy = cfg["page_strategy"]
    pages = select_pages(total, strategy)

    def _one(item):
        pno, img_url = item
        if cancel_check and cancel_check():
            return pno, None
        prompt = build_prompt(fields, pno, total)
        return pno, call_ocr(client, model_name, img_url,
                             prompt, cfg["max_tokens"])

    # Xử lý theo lô để giới hạn RAM: chiến lược 'full' trên PDF hàng nghìn trang
    # không nạp toàn bộ ảnh cùng lúc — mỗi lô chỉ giữ ~batch ảnh trong bộ nhớ.
    workers = max(1, min(cfg["concurrency"], len(pages)))
    batch_size = max(workers * 2, 4)

    collected, raws = [], {}
    sent_pages, sent_bytes = 0, 0        # thống kê ảnh base64 đã đẩy lên model
    for start in range(0, len(pages), batch_size):
        if cancel_check and cancel_check():
            break
        batch = pages[start:start + batch_size]
        rendered = render_pages(
            pdf_path, batch,
            cfg["render_dpi"], cfg["max_dimension"], cfg["jpeg_quality"],
        )
        if not rendered:
            continue
        sent_pages += len(rendered)
        for _pno, _url in rendered:
            b64 = _url.split(",", 1)[-1]
            sent_bytes += (len(b64) * 3) // 4     # độ dài base64 → số byte ảnh
        with ThreadPoolExecutor(max_workers=min(workers, len(rendered))) as ex:
            responses = list(ex.map(_one, rendered))
        for pno, raw in responses:
            if raw is None:
                continue
            raws[pno + 1] = raw
            parsed = parse_json_loose(raw)
            if parsed:
                collected.append((pno, parsed))

    if not collected:
        raise RuntimeError("không parse được JSON từ bất kỳ trang nào")

    merged, provenance = merge_page_results(
        collected, field_keys, cfg["field_page_preference"])

    audit = {}
    if corrector is not None:
        merged, audit = corrector.correct_fields(
            merged, cfg["correction_policy"])

    return {
        "file": str(pdf_path),
        "name": Path(pdf_path).name,
        "total_pages": total,
        "pages_processed": sorted(p + 1 for p, _ in collected),
        "sent_pages": sent_pages,
        "sent_bytes": sent_bytes,
        "data": merged,
        "provenance": provenance,
        "correction_audit": audit,
        "page_raw": raws,
    }


# ============================================================ GHI TSV

class TsvWriter:
    """Ghi liên tục sau từng file — crash giữa batch không mất dữ liệu."""

    def __init__(self, out_dir, field_keys):
        self.dir = Path(out_dir)
        self.dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%d.%m.%Y-%H.%M.%S")
        self.tsv_path = self.dir / f"ket_qua_{stamp}.tsv"
        self.jsonl_path = self.dir / f"ket_qua_{stamp}.jsonl"
        self.field_keys = list(field_keys)

        with open(self.tsv_path, "w", encoding="utf-8", newline="") as f:
            csv.writer(f, delimiter="\t").writerow(
                ["file", "timestamp"] + self.field_keys)

    @staticmethod
    def _clean(v):
        if v is None:
            return ""
        s = unicodedata.normalize("NFC", str(v))
        return " ".join(s.replace("\t", " ").split())

    def append(self, result: dict):
        ts = datetime.now().strftime("%d.%m.%Y-%H:%M:%S")
        row = [self._clean(result.get("name")), ts]
        data = result.get("data") or {}
        row += [self._clean(data.get(k, "")) for k in self.field_keys]

        with open(self.tsv_path, "a", encoding="utf-8", newline="") as f:
            csv.writer(f, delimiter="\t").writerow(row)

        with open(self.jsonl_path, "a", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False)
            f.write("\n")

    def append_error(self, name, message):
        with open(self.jsonl_path, "a", encoding="utf-8") as f:
            json.dump({"name": name, "status": "error", "error": message},
                      f, ensure_ascii=False)
            f.write("\n")


# ============================================================ XUẤT EXCEL

TEXT_FORMAT_COLUMNS = {"document_number", "issue_date"}


def excel_target(out_dir, source, base_name=None) -> Path:
    """
    Đường dẫn Excel mặc định, suy từ tên nguồn để dễ nhận ra file nào của
    thư mục nào. Cùng một nguồn chạy lại sẽ ra cùng đường dẫn — nghĩa là
    sẽ trùng, và đó là chỗ UI phải hỏi ghi đè.
    """
    out = Path(out_dir)
    if base_name:
        stem = base_name
    else:
        p = Path(source)
        stem = p.stem if p.suffix else (p.name or "ket_qua")
    return out / f"ket_qua_{stem}.xlsx"


def stamped_variant(target: Path) -> Path:
    """Biến thể có timestamp, dùng khi người dùng chọn giữ cả hai bản."""
    stamp = datetime.now().strftime("%d.%m.%Y-%H.%M.%S")
    return target.with_name(f"{target.stem}_{stamp}{target.suffix}")


def is_locked(target: Path) -> bool:
    """
    Đoán xem file có đang bị mở trong Excel không. Excel giữ khoá ghi nên
    mở chế độ append sẽ ném PermissionError. Kiểm tra trước rẻ hơn là để
    thất bại sau khi đã dựng xong workbook.
    """
    if not target.exists():
        return False
    try:
        with open(target, "a+b"):
            return False
    except (PermissionError, OSError):
        return True


def export_excel(tsv_path, target) -> Path:
    """Đọc TSV rồi ghi .xlsx vào đúng `target`. Ném PermissionError nếu bị khoá."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    tsv_path = Path(tsv_path)
    target = Path(target)
    target.parent.mkdir(parents=True, exist_ok=True)

    with open(tsv_path, "r", encoding="utf-8", newline="") as f:
        rows = list(csv.reader(f, delimiter="\t"))
    if not rows:
        raise RuntimeError("file TSV rỗng")

    header, body = rows[0], rows[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = "ket_qua"

    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    for r in body:
        ws.append([unicodedata.normalize("NFC", c) for c in r])

    # Cột mã và ngày phải là text, không thì Excel diễn giải lại theo locale
    for idx, name in enumerate(header, start=1):
        letter = get_column_letter(idx)
        if name in TEXT_FORMAT_COLUMNS:
            for row in range(2, len(body) + 2):
                ws[f"{letter}{row}"].number_format = "@"
        width = max([len(name)] + [len(r[idx - 1]) if idx - 1 < len(r) else 0
                                   for r in body[:200]])
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 55)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(target)
    return target
